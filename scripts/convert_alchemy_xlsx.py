"""Convert docs/data.xlsx alchemy sheets to runtime JSON data.

Reads these sheets and emits:
    Thảo Dược             → src/data/items/herbs.json
    Nguyên Liệu Yêu Thú   → src/data/items/yeu_thu.json
    Đan Dược              → src/data/items/pills.json
    Đan Phương            → src/data/recipes/pill_recipes.json

Run from the repo root:
    python scripts/convert_alchemy_xlsx.py

Section-header rows (marked with ✦ in any cell) are skipped. Recipe
ingredient cells may contain alternates separated by " / "; these are
parsed into a list of alternatives so alchemy.py can pick the first
satisfiable option (same pattern as forge recipes).
"""
from __future__ import annotations

import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl


REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "docs" / "data.xlsx"
OUT_HERBS = REPO_ROOT / "src" / "data" / "items" / "herbs.json"
OUT_YEUTHU = REPO_ROOT / "src" / "data" / "items" / "yeu_thu.json"
OUT_PILLS = REPO_ROOT / "src" / "data" / "items" / "pills.json"
OUT_RECIPES = REPO_ROOT / "src" / "data" / "recipes" / "pill_recipes.json"


# Quality chance table keyed by recipe grade (1-9). Higher grades have
# slightly more Thiên/Địa weighting to reward high-tier crafting.
QUALITY_CHANCES_BY_GRADE: dict[int, dict[str, float]] = {
    1: {"hoan": 0.78, "huyen": 0.18, "dia": 0.03, "thien": 0.01},
    2: {"hoan": 0.72, "huyen": 0.22, "dia": 0.05, "thien": 0.01},
    3: {"hoan": 0.66, "huyen": 0.25, "dia": 0.07, "thien": 0.02},
    4: {"hoan": 0.60, "huyen": 0.27, "dia": 0.10, "thien": 0.03},
    5: {"hoan": 0.54, "huyen": 0.28, "dia": 0.14, "thien": 0.04},
    6: {"hoan": 0.48, "huyen": 0.30, "dia": 0.17, "thien": 0.05},
    7: {"hoan": 0.42, "huyen": 0.30, "dia": 0.21, "thien": 0.07},
    8: {"hoan": 0.36, "huyen": 0.30, "dia": 0.25, "thien": 0.09},
    9: {"hoan": 0.30, "huyen": 0.30, "dia": 0.29, "thien": 0.11},
}


# Stack limits per item type.
STACK_HERB = 999
STACK_YEUTHU = 999
STACK_PILL = 99


# Map "Hiệu Ứng" free-text labels to internal effect keys. Unknown labels
# fall back to "misc_vi_label" so the raw text is preserved as ``vi_effect``.
EFFECT_KEY_MAP = {
    "Tăng EXP luyện thể":       "exp_luyen_the",
    "Tăng tu vi":                "exp_qi",
    "Hồi HP":                    "restore_hp",
    "Hồi MP":                    "restore_mp",
    "Tăng phòng thủ":            "buff_def",
    "Tăng tốc độ di chuyển":     "buff_speed",
    "Tăng sát thương kiếm pháp": "buff_sword_dmg",
    "Giảm đan độc tích lũy":     "reduce_toxicity",
    "Xua đuổi yêu thú":          "repel_beast",
    "Dụ yêu thú tiến đến":       "lure_beast",
    "Trúc cơ — hỗ trợ đột phá cảnh giới":        "breakthrough_truc_co",
    "Kết đan — đột phá lên Kim Đan cảnh":        "breakthrough_kim_dan",
    "Nguyên Anh — đột phá lên Nguyên Anh cảnh":  "breakthrough_nguyen_anh",
    "Hóa Thần cảnh — đột phá cảnh giới":         "breakthrough_hoa_than",
    "Luyện Hư cảnh — đột phá cảnh giới":         "breakthrough_luyen_hu",
    "Hợp Đạo cảnh — đột phá cảnh giới":          "breakthrough_hop_dao",
    "Đại Thừa cảnh — đột phá cảnh giới":         "breakthrough_dai_thua",
    "Tăng linh lực hệ Kim":  "buff_element_kim",
    "Tăng linh lực hệ Mộc":  "buff_element_moc",
    "Tăng linh lực hệ Thủy": "buff_element_thuy",
    "Tăng linh lực hệ Hỏa":  "buff_element_hoa",
    "Tăng linh lực hệ Thổ":  "buff_element_tho",
    "Tăng linh lực hệ Lôi":  "buff_element_loi",
    "Tăng linh lực hệ Phong":"buff_element_phong",
    "Tăng linh lực hệ Ám":   "buff_element_am",
}


def _is_section_header(row: list) -> bool:
    for cell in row:
        if isinstance(cell, str) and "✦" in cell:
            return True
    return False


def _strip(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_int(value, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default


# ── Sheet parsers ───────────────────────────────────────────────────────────

def parse_herbs(wb) -> list[dict]:
    ws = wb["Thảo Dược"]
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if _is_section_header(row):
            continue
        if not row[2]:  # key column empty → ignore
            continue
        key = _strip(row[2])
        out.append({
            "key": key,
            "vi": _strip(row[1]),
            "en": _strip(row[3]),
            "type": "herb",
            "grade": _parse_int(row[5], 1),
            "stack_max": STACK_HERB,
            "shop_price_merit": 0,
            "description_vi": _strip(row[6]),
        })
    return out


def parse_yeu_thu(wb) -> list[dict]:
    ws = wb["Nguyên Liệu Yêu Thú"]
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if _is_section_header(row):
            continue
        if not row[2]:
            continue
        out.append({
            "key": _strip(row[2]),
            "vi": _strip(row[1]),
            "en": _strip(row[3]),
            "type": "yeu_thu",
            "grade": _parse_int(row[5], 1),
            "stack_max": STACK_YEUTHU,
            "shop_price_merit": _parse_int(row[6], 0),
            "description_vi": _strip(row[7]),
            "source_rank": _strip(row[9]),
        })
    return out


def parse_pills(wb) -> list[dict]:
    ws = wb["Đan Dược"]
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if _is_section_header(row):
            continue
        if not row[2]:
            continue
        effect_vi = _strip(row[5])
        effect_key = EFFECT_KEY_MAP.get(effect_vi, "misc_vi_label")
        out.append({
            "key": _strip(row[2]),
            "vi": _strip(row[1]),
            "en": "",
            "type": "pill",
            "grade": _parse_int(row[3], 1),
            "stack_max": STACK_PILL,
            "shop_price_merit": 0,
            "category": _strip(row[4]),
            "effect_key": effect_key,
            "effect_vi": effect_vi,
            "dan_doc": _parse_int(row[6], 0),
            "description_vi": _strip(row[7]),
            "cultivation_gain": _strip(row[8]),
        })
    return out


_NAME_QTY_RE = re.compile(r"^(.*?)\s*(?:[xX×](\d+))\s*$")
_FURNACE_TIER_RE = re.compile(r"cấp\s*(\d+)", re.IGNORECASE)


def _parse_furnace_tier(text: str, recipe_grade: int) -> int:
    """Extract the furnace tier from a 'Lò luyện đan cấp N' string.

    The tier in the xlsx maxes out at 4 but recipes go up to grade 9, so a
    grade 9 pill still requires a tier-4 furnace. Fall back to clamp(grade,
    1, 4) when the text is missing or unparseable.
    """
    if text:
        m = _FURNACE_TIER_RE.search(text)
        if m:
            return max(1, min(4, int(m.group(1))))
    return max(1, min(4, recipe_grade))


def _resolve_herb_key(name: str, name_to_key: dict[str, str], missing: list[str]) -> str | None:
    name = name.strip()
    if not name:
        return None
    if name in name_to_key:
        return name_to_key[name]
    # Try case-insensitive / whitespace-normalised fallback
    norm = " ".join(name.split())
    for display, key in name_to_key.items():
        if " ".join(display.split()) == norm:
            return key
    missing.append(name)
    return None


def _parse_ingredient_cell(
    cell_value,
    sl_value,
    name_to_key: dict[str, str],
    missing: list[str],
) -> list[dict] | None:
    text = _strip(cell_value)
    if not text:
        return None
    # Alternates separated by " / "
    parts = [p.strip() for p in text.split("/")]
    options: list[dict] = []
    for part in parts:
        m = _NAME_QTY_RE.match(part)
        if m:
            name = m.group(1).strip()
            qty = int(m.group(2))
        else:
            name = part
            qty = _parse_int(sl_value, 1)
        key = _resolve_herb_key(name, name_to_key, missing)
        if key is None:
            continue
        options.append({"key": key, "qty": qty})
    return options or None


def parse_recipes(
    wb,
    pill_name_to_key: dict[str, str],
    ingredient_name_to_key: dict[str, str],
) -> tuple[list[dict], dict]:
    ws = wb["Đan Phương"]
    recipes: list[dict] = []
    missing_pills: list[str] = []
    missing_ingredients: list[str] = []
    skipped_no_grade = 0

    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if _is_section_header(row):
            continue
        if not row[2]:  # recipe ID empty → skip
            continue

        grade_raw = row[3]
        if not isinstance(grade_raw, (int, float)):
            skipped_no_grade += 1
            continue
        grade = int(grade_raw)

        pill_name = _strip(row[4])
        pill_key = pill_name_to_key.get(pill_name)
        if not pill_key:
            missing_pills.append(pill_name)
            continue

        chu = _parse_ingredient_cell(row[7], row[8], ingredient_name_to_key, missing_ingredients)
        phu = _parse_ingredient_cell(row[9], row[10], ingredient_name_to_key, missing_ingredients)
        dan = _parse_ingredient_cell(row[11], row[12], ingredient_name_to_key, missing_ingredients)

        ingredients: list[dict] = []
        for role, slot in (("chu", chu), ("phu", phu), ("dan", dan)):
            if slot:
                ingredients.append({"role": role, "options": slot})

        min_qi = max(0, min(8, grade - 1))

        # Cost scales exponentially with grade — mirrors forge economics.
        cost_cong_duc = 200 * (3 ** (grade - 1))

        furnace_text = _strip(row[5])
        furnace_tier = _parse_furnace_tier(furnace_text, grade)

        recipes.append({
            "key": _strip(row[2]),
            "vi": _strip(row[1]),
            "grade": grade,
            "min_qi_realm": min_qi,
            "output_pill": pill_key,
            "furnace": furnace_text,
            "furnace_tier": furnace_tier,
            "craft_time_display": _strip(row[6]),
            "cost_cong_duc": cost_cong_duc,
            "ingredients": ingredients,
            "quality_chances": QUALITY_CHANCES_BY_GRADE.get(grade, QUALITY_CHANCES_BY_GRADE[9]),
        })

    stats = {
        "count": len(recipes),
        "skipped_no_grade": skipped_no_grade,
        "missing_pills": list(OrderedDict.fromkeys(missing_pills)),
        "missing_ingredients": list(OrderedDict.fromkeys(missing_ingredients)),
    }
    return recipes, stats


# ── Writer ──────────────────────────────────────────────────────────────────

def _dedupe_by_key(entries: list[dict]) -> list[dict]:
    """Keep the last entry per ``key`` — xlsx sometimes repeats a key across
    cross-referenced sections; registry would silently override anyway."""
    seen: dict[str, dict] = {}
    for entry in entries:
        seen[entry["key"]] = entry
    return list(seen.values())


def _write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found", file=sys.stderr)
        return 1

    print(f"Loading {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    herbs = _dedupe_by_key(parse_herbs(wb))
    yeu_thu = _dedupe_by_key(parse_yeu_thu(wb))
    pills = _dedupe_by_key(parse_pills(wb))

    ingredient_name_to_key: dict[str, str] = {}
    for entry in herbs + yeu_thu:
        ingredient_name_to_key[entry["vi"]] = entry["key"]
    pill_name_to_key = {p["vi"]: p["key"] for p in pills}

    recipes, stats = parse_recipes(wb, pill_name_to_key, ingredient_name_to_key)
    recipes = _dedupe_by_key(recipes)

    _write_json(OUT_HERBS, herbs)
    _write_json(OUT_YEUTHU, yeu_thu)
    _write_json(OUT_PILLS, pills)
    _write_json(OUT_RECIPES, recipes)

    print(f"  herbs:    {len(herbs):3d} → {OUT_HERBS.relative_to(REPO_ROOT)}")
    print(f"  yeu_thu:  {len(yeu_thu):3d} → {OUT_YEUTHU.relative_to(REPO_ROOT)}")
    print(f"  pills:    {len(pills):3d} → {OUT_PILLS.relative_to(REPO_ROOT)}")
    print(f"  recipes:  {stats['count']:3d} → {OUT_RECIPES.relative_to(REPO_ROOT)}")
    if stats["skipped_no_grade"]:
        print(f"  [warn] skipped {stats['skipped_no_grade']} recipe rows without numeric grade")
    if stats["missing_pills"]:
        print(f"  [warn] {len(stats['missing_pills'])} recipes reference unknown pill names:")
        for name in stats["missing_pills"][:10]:
            print(f"         - {name}")
    if stats["missing_ingredients"]:
        print(f"  [warn] {len(stats['missing_ingredients'])} unknown ingredient names (not in Thảo Dược / Nguyên Liệu Yêu Thú):")
        for name in stats["missing_ingredients"][:10]:
            print(f"         - {name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
